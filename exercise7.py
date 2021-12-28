#把结果写入excel
# import openpyxl
# sc = openpyxl.load_workbook('test_case_api.xlsx')
# sheet =sc['register']
# sheet.cell(row=2,column=8).value='pass'
# sc.save('test_case_api.xlsx')


#写入程序函数化
import openpyxl
def write(filename,sheetname,row,column,final_result):
    sc = openpyxl.load_workbook(filename)
    sheet =sc[sheetname]
    sheet.cell(row=row,column=column).value= final_result
    sc.save(filename)


#自动化流程化
#读取函数化
import requests
def read_data(file_name,sheet_name):
    excel=openpyxl.load_workbook( file_name)
    sheet2=excel[sheet_name]
    max_row =sheet2.max_row
    list1=[]                                #定义一个空列表  用来接收测试数据
    for i in range(2,max_row+1):
        dict1=dict(
        id=sheet2.cell(row=i,column=1).value ,   #取id
        url= sheet2.cell(row=i,column=5).value,   #去url
        data=sheet2.cell(row=i,column=6).value,    #去data
        expect=sheet2.cell(row=i,column=7).value  #取expect期望值
        )
        list1.append(dict1)                    #将字典里的数据加到列表里-->列表就存放了所有测试数据
    return list1

#发送请求函数
def api_func(url, res_body):
    requests_header = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    res = requests.post(url=url, json=res_body, headers=requests_header)
    res_log = res.json()
    return res_log

read = read_data('test_case_api.xlsx','register')
for testcase in read:
    case_id = testcase.get('id')
    url = testcase.get('url')
    data = testcase.get('data')   #取出的data从excel取出数据类型为str
    data = eval(data)   #去除引号‘’
    expect =testcase.get('expect')
    expect=eval(expect)
    expect_msg = expect.get('msg')
    app=api_func(url=url,res_body=data)
    # print(app)
    real_msg=app.get('msg')
    print('预期结果为:{}'.format(expect_msg))
    print('实际结果为：{}'.format(real_msg))
    if real_msg == expect_msg:
        print('pass')
        final_res='pass'
    else:
        print('false')
        final_res='false'
    print('*'*100)
    write('test_case_api.xlsx','register',case_id+1,8,final_res)
