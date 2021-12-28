#接口自动化的步骤
#使用excel整理好接口测试用例
#通过python去读取excel的测试数据
#使用requests去发送请求
#断言（判断）：执行结果 vs 预期结果
#excel填写结果： 通过/不通过

#读数据（excel）    #需要安装openpyxl第三方库
import openpyxl
# requests     发送hhttp
# jsonpath     提取数据
# openpyxl     读，写execl

# 操作excel三大对象
# 找到工作簿excel    把excel放到package下
# 找到表格sheet
# 找单元格cell

#找工作簿openpyxl.load_workbook('')
# excel=openpyxl.load_workbook('test_case_api.xlsx')    #将excel加载到python的内存中
# #找到表单
# sheet=excel['login']
# #找到单元格
# #row行  colum列  .cell为固定形式
# cell1=sheet.cell(row=2,column=2)
# #表示单元格
# print(cell1)
# #表示单元格内容
# print(cell1.value)
#
# #写入数据
# cell1.value='spirit'
# #若不保存则不能修改execl，并且源文件需要关闭才行
# excel.save('test_case_api.xlsx')

#自动化的读取excel里需要的接口测试数据
# excel=openpyxl.load_workbook('test_case_api.xlsx')
# sheet2=excel['login']
# max_row =sheet2.max_row
# list1=[]                                #定义一个空列表  用来接收测试数据
# for i in range(2,max_row+1):
#     dict1=dict(
#     id=sheet2.cell(row=i,column=1).value ,   #取id
#     url= sheet2.cell(row=i,column=5).value,   #去url
#     data=sheet2.cell(row=i,column=6).value,    #去data
#     expect=sheet2.cell(row=i,column=7).value  #取expect期望值
#     )
#     list1.append(dict1)                    #将字典里的数据加到列表里-->列表就存放了所有测试数据
# print(list1)

#函数化
# def read_data(file_name,sheet_name):
#     excel=openpyxl.load_workbook( file_name)
#     sheet2=excel[sheet_name]
#     max_row =sheet2.max_row
#     list1=[]                                #定义一个空列表  用来接收测试数据
#     for i in range(2,max_row+1):
#         dict1=dict(
#         id=sheet2.cell(row=i,column=1).value ,   #取id
#         url= sheet2.cell(row=i,column=5).value,   #去url
#         data=sheet2.cell(row=i,column=6).value,    #去data
#         expect=sheet2.cell(row=i,column=7).value  #取expect期望值
#         )
#         list1.append(dict1)                    #将字典里的数据加到列表里-->列表就存放了所有测试数据
#     return list1
#
# rest=read_data('test_case_api.xlsx','login')
# print(rest)
#买水果
#买10个西瓜，10个榴莲，10个香蕉
#售货员打包，分类的打包，10个西瓜放一起，10个留恋放一起，10个香蕉放一起
#最后再用一个大袋子，再把三个小袋子装到一起





