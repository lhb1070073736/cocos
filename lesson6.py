# -*- coding: utf-8 -*-
# @Author   :   YaMeng
# @File :   lesson6.py
# @Software :   PyCharm
# @Time :   2020/8/21 9:33
# @company  :   湖南省零檬信息技术有限公司

# 接口自动化的步骤
# 使用excel整理好接口测试用例   -- done
# 通过python去读取excel的测试数据  -- done
# 使用requests去发送请求，并且得到响应结果  -- done
# 断言：执行结果  vs  预期结果
# 把通过/不通过的最终结果要回写到excel


# 第三方库
# requests -- 发送http请求
# jsonpath -- 提取数据
# openpyxl -- 读、写excel

# 操作excel三大对象
# 找到工作簿： excecl表格   -- 1、把excel放到package下面，好处就是不用频繁去写路径 2、每次用的时候都带上路径
# 找到表单： sheet
# 找单元格：cell单元格

import  openpyxl


# # 读取数据
# # 找到工作簿
# biaoge = openpyxl.load_workbook('test_case_api.xlsx')  # 将excel加载到python的内存中
# # 找到表单：sheet
# sheet = biaoge['register']
# # # 找到单元格
# # row 行  column 列
# # .cell是取单元格的方法，是固定
# danyuange = sheet.cell(row=2, column=3)
# print(danyuange)   # 取的是单元格，并不是单元格内的数据
# print(danyuange.value)  # 加上.value 才会取的是单元格内的数据

#
# # 写入数据
# danyuange.value = 'spirit'
# # 写入完了之后，要保存excel才是真正的写入了。没有保存那excel里是没有的
# # 从pacharm里打开的excel，一定是要关掉后再执行代码保存。不然就会提示操作被拒绝
# biaoge.save('test_case_api.xlsx')


# 自动化的读取excel里需要的接口测试数据
import openpyxl

# def read_data(filename, sheetname):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[sheetname]
#     max_row = sheet.max_row  # 获取sheet里最大的行数
#     case_list = []    # 定义一个空列表，是专门用来接收测试数据的
#     for i in range(2, max_row+1):  # 因为取头不取尾，所以要在最大行数的基础上 +1
#         dict1 = dict(   #  转换成字典就是为了将测试用例打包成一条一条的数据
#         id = sheet.cell(row=i, column=1).value,        #  取的是测试用来编号id
#         url = sheet.cell(row=i, column=5).value,       #  取的是url数据
#         data = sheet.cell(row=i, column=6).value,      #  取的是data数据
#         expect = sheet.cell(row=i, column=7).value)    #  取的是预期结果（expected数据）
#         case_list.append(dict1)     # 用append把字典追加到列表去  --> 列表就存放了所有的测试数据
#     # print(case_list)
#     return case_list   # 设置返回值，给别人去用
# lll = read_data('test_case_api.xlsx','login')
# print(lll)
# print(case_list[2])

# 去买水果
# 买 10个西瓜。10个榴莲，10个香蕉
# 售货员打包，分类的打包，10个西瓜放一起，袋子装好。 10个榴莲装一下，10个香蕉装一起
# 最后再用一个大的袋子，再把这3个小袋子给装一起去。

# 到家之后
# 先把最大的袋子去掉。把里面的小袋子拿出来
# 再一个一个的把西瓜或者是榴莲或者是香蕉给从小袋子再拿出来

# res = read_data('test_case_api.xlsx', 'login')
# print(res)